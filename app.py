# -*- coding: utf-8 -*-
import os
import sys
# スクレイピング用
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
# ロギング用
import logging
import logzero
from logzero import logger
# スリープ用
import time
# 設定ファイル読み込み&出力
import configparser
# ディレクトリ、ファイルパス取得用
import tkinter, tkinter.filedialog
# コマンドラインパーサ
import click
# 現在日付取得用
from datetime import datetime
# Excelファイル操作(Excelファイルに画像を貼り付ける用にPillowもインストールが必要、インポートは不要)
import openpyxl
# 暗号化/複合化を行うクラス
from encrypter import simple_encrypter
# バックアップ用
import shutil

# 定数
CONFIG_FILE_NAME = 'settings.ini'
CONFIG_OPT_LOGIN_ID = 'login_id'
CONFIG_OPT_PASSWORD = 'password'
CONFIG_OPT_CHROME_EXECUTABLE_PATH = 'chrome_executable_path'
CONFIG_OPT_TRADE_PERFORMANCE_XLSX_PATH = 'trade_performance_xlsx_path'
CURRENT_DATE = datetime.now().strftime("%Y%m%d")
WORK_DIR = 'work' + os.sep + CURRENT_DATE + os.sep
LOG_DIR = 'log' + os.sep
LOG_FILE = LOG_DIR + 'application.log'
ENCRYPTION_KEY = 'UWAm1mweGbaCdwab'

# ログファイル
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
logzero.logfile(LOG_FILE, encoding = "utf-8")
logzero.loglevel(logging.INFO)

def get_config():
    """
    設定ファイルから設定を取得する
    設定ファイルに設定が存在しない場合は対話形式で設定を作成する
    """
    config_name_default = 'default'
    config_name_chrome = 'chrome'

    config = configparser.ConfigParser()
    config.read(CONFIG_FILE_NAME, 'cp932')

    # 設定ファイル内にchrome,defaultセクションが存在しない場合は作成
    if not config.has_section(config_name_chrome):
        logger.debug("{0}に{1}セクションを作成します".format(CONFIG_FILE_NAME, config_name_chrome))
        config.add_section(config_name_chrome)
    if not config.has_section(config_name_default):
        logger.debug("{0}に{1}セクションを作成します".format(CONFIG_FILE_NAME, config_name_default))
        config.add_section(config_name_default)

    # ChromeDriverのパス
    if not config.has_option(config_name_chrome, CONFIG_OPT_CHROME_EXECUTABLE_PATH):
        logger.info("ChromeDriverのパスが設定されていないので、設定してください")
        root = tkinter.Tk()
        root.withdraw()
        chrome_executable_path = tkinter.filedialog.askopenfilename(filetypes = [("ChromeDriverの実行ファイル", "*.exe")], initialdir = os.getcwd())
        if chrome_executable_path == '':
            logger.error("ChromeDriverのパスは必須です")
            sys.exit(1)
        logger.debug("選択されたファイル：{0}".format(chrome_executable_path))
        config.set(config_name_chrome, CONFIG_OPT_CHROME_EXECUTABLE_PATH, chrome_executable_path)
        with open(CONFIG_FILE_NAME, 'w') as config_file:
            logger.debug("{0}の{1}セクションに{2}を追記して保存します".format(CONFIG_FILE_NAME, config_name_chrome, CONFIG_OPT_CHROME_EXECUTABLE_PATH))
            config.write(config_file)
    else:
        chrome_executable_path = config.get(config_name_chrome, CONFIG_OPT_CHROME_EXECUTABLE_PATH)

    # ログインID
    if not config.has_option(config_name_default, CONFIG_OPT_LOGIN_ID):
        logger.info("SBI証券のログインIDを入力してください")
        login_id = input("入力：")
        if login_id == '':
            logger.error("SBI証券のログインIDは必須です")
            sys.exit(1)
        logger.debug("入力されたログインID：{0}".format(login_id))
        saving_login_id = simple_encrypter.encrypt(login_id, ENCRYPTION_KEY)
        config.set(config_name_default, CONFIG_OPT_LOGIN_ID, saving_login_id)
        with open(CONFIG_FILE_NAME, 'w') as config_file:
            logger.debug("{0}の{1}セクションに{2}を追記して保存します".format(CONFIG_FILE_NAME, config_name_default, CONFIG_OPT_LOGIN_ID))
            config.write(config_file)
    else:
        login_id = config.get(config_name_default, CONFIG_OPT_LOGIN_ID)
        login_id = simple_encrypter.decrypt(login_id, ENCRYPTION_KEY)

    # ログインパスワード(パスワードはsettings.iniに保存する際に暗号化する)
    if not config.has_option(config_name_default, CONFIG_OPT_PASSWORD):
        logger.info("SBI証券のログインパスワードを入力してください")
        password = input("入力：")
        if password == '':
            logger.error("SBI証券のログインパスワードは必須です")
            sys.exit(1)
        saving_password = simple_encrypter.encrypt(password, ENCRYPTION_KEY)
        config.set(config_name_default, CONFIG_OPT_PASSWORD, saving_password)
        with open(CONFIG_FILE_NAME, 'w') as config_file:
            logger.debug("{0}の{1}セクションに{2}を追記して保存します".format(CONFIG_FILE_NAME, config_name_default, CONFIG_OPT_PASSWORD))
            config.write(config_file)
    else:
        password = config.get(config_name_default, CONFIG_OPT_PASSWORD)
        password = simple_encrypter.decrypt(password, ENCRYPTION_KEY)

    # Trade-Performance-2021(xlsx)のパス
    if not config.has_option(config_name_default, CONFIG_OPT_TRADE_PERFORMANCE_XLSX_PATH):
        logger.info("Trade-Performance-2021.xlsxのパスが設定されていないので、設定してください")
        root = tkinter.Tk()
        root.withdraw()
        trade_performance_xlsx_path = tkinter.filedialog.askopenfilename(filetypes = [("Trade-Performance-2021.xlsx", "*.xlsx")], initialdir = os.getcwd())
        if trade_performance_xlsx_path == '':
            logger.error("Trade-Performance-2021.xlsxのパスは必須です")
            sys.exit(1)
        logger.debug("選択されたファイル：{0}".format(trade_performance_xlsx_path))
        config.set(config_name_default, CONFIG_OPT_TRADE_PERFORMANCE_XLSX_PATH, trade_performance_xlsx_path)
        with open(CONFIG_FILE_NAME, 'w') as config_file:
            logger.debug("{0}の{1}セクションに{2}を追記して保存します".format(CONFIG_FILE_NAME, config_name_default, CONFIG_OPT_TRADE_PERFORMANCE_XLSX_PATH))
            config.write(config_file)
    else:
        trade_performance_xlsx_path = config.get(config_name_default, CONFIG_OPT_TRADE_PERFORMANCE_XLSX_PATH)

    return {
        'chrome_executable_path': chrome_executable_path,
        'login_id': login_id,
        'password': password,
        'trade_performance_xlsx_path': trade_performance_xlsx_path,
    }

def save_current_html_source(driver, debug_log_title, htmlname):
    """
    seleniumが参照中のhtmlソースを保存する
    """
    logger.debug(debug_log_title)
    with open(WORK_DIR + htmlname, 'w', encoding='utf-8') as f:
        f.write(driver.page_source)

# 設定取得
config = get_config()

@click.command(context_settings = dict(help_option_names = ['-h', '--help']))
@click.option('--debug', is_flag = True, help = "debugログを出力します")
def main(debug):

    logger.info("trade-performance-auto-input-2021 start.")
    if debug:
        logzero.loglevel(logging.DEBUG)

    logger.info("workディレクトリに本日分の作業フォルダ作成")
    if not os.path.exists(WORK_DIR):
        os.makedirs(WORK_DIR)

    logger.info("Chromeを起動")
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/88.0.4324.104 Safari/537.36')
    options.add_argument('--guest')
    if not debug:
        options.add_argument('--headless')

    driver = webdriver.Chrome(executable_path = config['chrome_executable_path'], options = options)

    logger.info("SBI証券にログイン")
    logger.debug("ログイン画面を開く")
    driver.get("https://www.sbisec.co.jp/ETGate")

    logger.debug("ログイン情報を入力、ログインボタンクリック")
    driver.find_element_by_css_selector("input[name='user_id']").send_keys(config['login_id'])
    driver.find_element_by_css_selector("input[name='user_password']").send_keys(config['password'])
    driver.find_element_by_css_selector("input[name='ACT_login']").click()
    if debug:
        save_current_html_source(driver, 'ログイン後画面のソースを保存', 'login_after.html')

    logger.info("口座管理画面を開く")
    driver.get('https://site2.sbisec.co.jp/ETGate/?_ControlID=WPLETacR001Control&_PageID=DefaultPID&_DataStoreID=DSWPLETacR001Control&_ActionID=DefaultAID&getFlg=on')
    if debug:
        save_current_html_source(driver, '口座管理画面のソースを保存', 'account_management.html')

    logger.debug('口座管理画面の計欄の数値を取得')
    sum_selector_path = 'body > div:nth-child(1) > table > tbody > tr > td:nth-child(1) > table > tbody > tr:nth-child(2) > td > table:nth-child(1) > tbody > tr > td > form > table:nth-child(3) > tbody > tr:nth-child(1) > td:nth-child(2) > table:nth-child(20) > tbody > tr > td:nth-child(1) > table:nth-child(7) > tbody > tr:nth-child(8) > td:nth-child(2) > div > b'
    current_sum = driver.find_element_by_css_selector(sum_selector_path).text

    logger.info("GoogleChrome正常終了")
    driver.close()
    driver.quit()

    logger.info("バックアップを作成してExcelに書き込む")
    shutil.copy(config['trade_performance_xlsx_path'], WORK_DIR + os.path.basename(config['trade_performance_xlsx_path']))
    today_md_slash = datetime.today().strftime('%#m/%#d')
    wb = openpyxl.load_workbook(config['trade_performance_xlsx_path'])

    target_sheet_name = datetime.today().strftime('%#m') + '月'
    logger.debug(target_sheet_name + 'のシート取得')
    ws = wb[target_sheet_name]

    # A4～A30の範囲の月日と今日の月日を比較して計の書き込み先を見つける
    target_row_num = None
    for row in ws['A4:A26']: #一番営業日が多い月(3月)に揃える
        for cell in row:
            if (today_md_slash == openpyxl.utils.datetime.from_excel(cell.value).strftime('%#m/%#d')):
                logger.debug('A列に今日の日付が見つかりました。見つけた日付：{0}'.format(today_md_slash))
                target_row_num = cell.row
                break
        else:
            continue
        break
    else:
        logger.debug('A列に今日の日付が見つかりませんでした。')
        sys.exit(1)

    # 入力シート内の追記位置が見つかったら書き込んで保存
    if target_row_num is not None:
        # M列(口座A)に上で取得した計を追記
        ws['M' + str(target_row_num)] = current_sum
        wb.save(config['trade_performance_xlsx_path'])

    logger.info("trade-performance-auto-input-2021 end.")
    return


if __name__ == "__main__":
    main()